#!/usr/bin/env python3
"""
交易查詢工具 - 查詢區塊鏈地址的交易數量
支持代理輪換功能以繞過RPC IP限制
"""

import requests
import json
import random
import time
import sys
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from utils.console import (
    COLOR_YELLOW,
    color_text,
    print_error,
    print_success,
    print_warn,
)

try:
    import tomllib
except ImportError:
    import tomli as tomllib


def load_config(config_file: str = 'config.toml') -> Dict:
    """
    加載TOML配置文件
    
    Args:
        config_file: 配置文件路徑
        
    Returns:
        配置字典
    """
    try:
        with open(config_file, 'rb') as f:
            return tomllib.load(f)
    except FileNotFoundError:
        print(f"✗ 找不到配置文件: {config_file}")
        sys.exit(1)
    except Exception as e:
        print(f"✗ 讀取配置文件出錯: {str(e)}")
        sys.exit(1)


def get_enabled_network(config: Dict) -> str:
    """
    獲取第一個啟用的網絡名稱
    
    Args:
        config: 配置字典
        
    Returns:
        啟用的網絡名稱
    """
    networks = config.get('networks', {})
    for network_name, network_config in networks.items():
        if network_config.get('enabled', False):
            return network_name
    raise ValueError("config.toml 中沒有啟用任何網絡，請設定 enabled = true")


config = load_config()
settings = config.get('settings', {})

SELECTED_NETWORK = get_enabled_network(config)
BATCH_SIZE = settings['batch_size']
BATCH_DELAY = settings['batch_delay']
CONCURRENT_REQUESTS = settings['concurrent_requests']
USE_PROXY = settings['use_proxy']
PROXY_FILE = settings['proxy_file']
PROXY_ROTATION_STRATEGY = settings['proxy_rotation_strategy']
PROXY_TIMEOUT = settings['proxy_timeout']
PROXY_RETRY_TIMES = settings['proxy_retry_times']
PROXY_FALLBACK_NO_PROXY = settings['proxy_fallback_no_proxy']
OUTPUT_FORMATS = settings['output_formats']
IMPORT_FILE = settings['import_file']
DEBUG_MODE = settings.get('debug', False)
RETRY_DELAY_SECONDS = settings.get('retry_delay_seconds', 0.0)
RETRY_FAILED_WALLETS = settings.get('retry_failed_wallets', False)
RETRY_UNTIL_ALL_SUCCESS = settings.get('retry_until_all_success', False)
MAX_RETRY_ROUNDS = settings.get('max_retry_rounds', 1)
RETRY_ROUND_DELAY = settings.get('retry_round_delay', 0.0)


def ensure_file_with_sample(target_file: str) -> bool:
    """
    若目標文件不存在，詢問是否用對應 -Sample 文件複製建立

    Args:
        target_file: 目標文件路徑

    Returns:
        是否可用（存在或已成功建立）
    """
    target_path = Path(target_file)
    if target_path.exists():
        return True

    sample_path = target_path.with_name(f"{target_path.stem}-Sample{target_path.suffix}")
    print_warn(f"⚠️  找不到文件: {target_path}")

    if not sample_path.exists():
        print_error(f"✗ 也找不到樣本文件: {sample_path}")
        return False

    while True:
        answer = input(f"是否使用樣本文件 {sample_path.name} 建立 {target_path.name}？(y/n): ").strip().lower()
        if answer in ("y", "yes"):
            try:
                shutil.copyfile(sample_path, target_path)
                print_success(f"✓ 已建立: {target_path}")
                return True
            except Exception as e:
                print_error(f"✗ 建立文件失敗: {e}")
                return False
        if answer in ("n", "no"):
            print_error(f"✗ 已取消建立 {target_path.name}")
            return False
        print("請輸入 y 或 n")


class ProxyManager:
    """代理管理類"""
    
    def __init__(self, proxy_file: str = PROXY_FILE, strategy: str = PROXY_ROTATION_STRATEGY):
        """
        初始化代理管理器
        
        Args:
            proxy_file: 代理文件路徑
            strategy: 輪換策略 ('rotate' 或 'random')
        """
        self.proxy_file = proxy_file
        self.strategy = strategy
        self.proxies = []
        self.current_index = 0
        self.load_proxies()
    
    def load_proxies(self) -> bool:
        """
        從文件加載代理列表
        格式: IP:Port:Username:Password
        
        Returns:
            是否成功加載
        """
        self.proxies = []
        try:
            with open(self.proxy_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    # 跳過空行和注釋
                    if line and not line.startswith('#'):
                        self.proxies.append(line)
            
            if self.proxies:
                print(f"✓ 已加載 {len(self.proxies)} 個代理\n")
                return True
            else:
                print(f"⚠️  代理文件為空: {self.proxy_file}\n")
                return False
        
        except FileNotFoundError:
            print(f"✗ 找不到代理文件: {self.proxy_file}\n")
            return False
        except Exception as e:
            print(f"✗ 讀取代理文件出錯: {str(e)}\n")
            return False
    
    @staticmethod
    def parse_proxy(proxy_string: str) -> Optional[Dict[str, str]]:
        """
        解析代理字符串為requests格式
        格式: IP:Port:Username:Password
        
        Args:
            proxy_string: 代理字符串
            
        Returns:
            {'https': 'https://user:pass@ip:port'} 或 None
        """
        try:
            parts = proxy_string.split(':')
            if len(parts) != 4:
                return None
            
            ip, port, username, password = parts
            proxy_url = f"https://{username}:{password}@{ip}:{port}"
            
            return {
                'https': proxy_url,
                'http': proxy_url
            }
        except Exception as e:
            print(f"  ⚠️  代理解析失敗 ({proxy_string}): {str(e)}")
            return None
    
    def get_next_proxy(self) -> Optional[Dict[str, str]]:
        """
        獲取下一個代理
        
        Returns:
            代理字典或None
        """
        if not self.proxies:
            return None
        
        if self.strategy == 'random':
            proxy_string = random.choice(self.proxies)
        else:  # rotate
            proxy_string = self.proxies[self.current_index]
            self.current_index = (self.current_index + 1) % len(self.proxies)
        
        return self.parse_proxy(proxy_string)


class TxChecker:
    """區塊鏈交易查詢類"""
    
    def __init__(self, network: str = SELECTED_NETWORK, use_proxy: bool = USE_PROXY):
        """
        初始化查詢器
        
        Args:
            network: 選擇的網絡名稱
            use_proxy: 是否使用代理
        """
        self.network = network
        self.use_proxy = use_proxy
        self.proxy_manager = None
        self.request_count = 0
        self.proxy_error_count = 0
        
        networks_config = config.get('networks', {})
        if network in networks_config:
            rpc_list = networks_config[network].get('rpc', [])
            if rpc_list:
                self.rpc_url = random.choice(rpc_list)
                self.network_info = {'name': network, 'chain_id': 'unknown'}
            else:
                raise ValueError(f"網絡 '{network}' 沒有配置RPC端點")
        else:
            raise ValueError(f"不支持的網絡: {network}. 支持的網絡: {list(networks_config.keys())}")
        
        # 初始化代理管理器
        if self.use_proxy:
            self.proxy_manager = ProxyManager()
            if not self.proxy_manager.proxies:
                print("⚠️  未找到可用代理，將使用直接連接\n")
                self.use_proxy = False
        
        print(f"✓ 已連接到: {self.network_info['name']}")
        print(f"✓ RPC URL: {self.rpc_url}")
        if self.use_proxy and self.proxy_manager and self.proxy_manager.proxies:
            print(f"✓ 代理模式: {PROXY_ROTATION_STRATEGY} (共 {len(self.proxy_manager.proxies)} 個代理)")
        print()

    def _short_error_message(self, error_msg: str) -> str:
        """將錯誤訊息縮短為一般使用者可讀版本"""
        msg = (error_msg or "").replace("\n", " ").strip()
        lowered = msg.lower()

        if "header not found" in lowered or "missing data" in lowered:
            return "RPC節點資料不完整（header not found / missing data）"
        if "connection refused" in lowered:
            return "RPC上游拒絕連線（connection refused）"
        if "cannot parse json-rpc response" in lowered or "parseerror" in lowered:
            return "RPC回傳格式異常（非標準JSON-RPC）"
        if "timeout" in lowered or "超時" in lowered:
            return "請求超時"
        if "proxy" in lowered or "代理" in lowered:
            return "代理連線異常"

        return msg[:120] + ("..." if len(msg) > 120 else "")

    def _format_error(self, error_msg: str) -> str:
        """依 debug 模式決定顯示完整或精簡錯誤"""
        if DEBUG_MODE:
            return error_msg
        return self._short_error_message(error_msg)
    
    def get_transaction_count(self, address: str) -> Tuple[Optional[int], Optional[str]]:
        """
        查詢地址的交易數量
        
        Args:
            address: 以太坊地址
            
        Returns:
            (交易數量, 錯誤訊息)；成功時錯誤訊息為None
        """
        # 標準化地址
        if not address.startswith('0x'):
            address = '0x' + address
        
        payload = {
            "jsonrpc": "2.0",
            "method": "eth_getTransactionCount",
            "params": [address, "latest"],
            "id": 1
        }
        
        last_error = None
        
        # 重試邏輯
        for attempt in range(PROXY_RETRY_TIMES):
            try:
                proxies = None
                proxy_info = ""
                
                # 獲取代理
                if self.use_proxy and self.proxy_manager:
                    proxies = self.proxy_manager.get_next_proxy()
                    if proxies:
                        proxy_info = f" [代理 {self.proxy_manager.current_index}/{len(self.proxy_manager.proxies)}]"
                
                self.request_count += 1
                
                response = requests.post(
                    self.rpc_url,
                    json=payload,
                    timeout=PROXY_TIMEOUT if self.use_proxy else 10,
                    proxies=proxies if proxies else None,
                    verify=False  # 對於HTTPS代理可能需要跳過SSL驗證
                )
                response.raise_for_status()
                
                result = response.json()
                
                if 'error' in result:
                    error_msg = result['error'].get('message', '未知錯誤')
                    last_error = f"錯誤: {self._format_error(error_msg)}"
                    return None, last_error
                
                # 將十六進制轉換為十進制
                tx_count = int(result['result'], 16)
                return tx_count, None
                
            except requests.exceptions.ProxyError as e:
                self.proxy_error_count += 1
                last_error = f"代理錯誤: {self._format_error(str(e))}"
                if attempt < PROXY_RETRY_TIMES - 1:
                    if RETRY_DELAY_SECONDS > 0:
                        time.sleep(RETRY_DELAY_SECONDS)
                    continue
                elif PROXY_FALLBACK_NO_PROXY:
                    try:
                        response = requests.post(
                            self.rpc_url,
                            json=payload,
                            timeout=10,
                            proxies=None
                        )
                        response.raise_for_status()
                        result = response.json()
                        if 'error' in result:
                            last_error = f"錯誤: {self._format_error(result['error'].get('message', '未知錯誤'))}"
                            return None, last_error
                        tx_count = int(result['result'], 16)
                        return tx_count, None
                    except Exception as fallback_error:
                        last_error = f"直接連接失敗: {self._format_error(str(fallback_error))}"
                        return None, last_error
                else:
                    return None, last_error
            
            except requests.exceptions.Timeout:
                last_error = f"超時: {self._format_error('連接超時')}"
                if attempt < PROXY_RETRY_TIMES - 1:
                    if RETRY_DELAY_SECONDS > 0:
                        time.sleep(RETRY_DELAY_SECONDS)
                    continue
                else:
                    return None, last_error
            
            except requests.exceptions.RequestException as e:
                last_error = f"網絡錯誤: {self._format_error(str(e))}"
                if attempt < PROXY_RETRY_TIMES - 1:
                    if RETRY_DELAY_SECONDS > 0:
                        time.sleep(RETRY_DELAY_SECONDS)
                    continue
                else:
                    return None, last_error
            
            except (json.JSONDecodeError, KeyError, ValueError) as e:
                last_error = f"解析錯誤: {self._format_error(str(e))}"
                return None, last_error
        
        return None, last_error
    
    def check_multiple_addresses(self, addresses: List[str]) -> Dict[str, int]:
        """
        查詢多個地址的交易數量 - 支持分批並行查詢
        
        Args:
            addresses: 地址列表
            
        Returns:
            {地址: 交易數量} 的字典
        """
        # 過濾空行和注釋
        valid_addresses = [
            addr.strip() for addr in addresses 
            if addr.strip() and not addr.strip().startswith('#')
        ]
        
        all_results: Dict[str, int] = {}
        pending_addresses = valid_addresses[:]
        round_num = 1
        total_input = len(valid_addresses)

        print(f"正在查詢 {total_input} 個地址的交易數量...")
        print(f"批次大小: {BATCH_SIZE}, 並行查詢數: {CONCURRENT_REQUESTS}, 批次延遲: {BATCH_DELAY}s")
        if RETRY_FAILED_WALLETS:
            rounds_text = "無限" if MAX_RETRY_ROUNDS == 0 else str(MAX_RETRY_ROUNDS)
            print(f"失敗重試: 已啟用（重試間隔 {RETRY_DELAY_SECONDS}s，每輪間隔 {RETRY_ROUND_DELAY}s，最大輪數 {rounds_text}）")
        print()

        while pending_addresses:
            if MAX_RETRY_ROUNDS > 0 and round_num > MAX_RETRY_ROUNDS:
                print_warn(f"⚠️  已達最大輪數 {MAX_RETRY_ROUNDS}，停止重試。")
                break

            total_addresses = len(pending_addresses)
            round_failures: List[str] = []

            print(f"\n{color_text(f'===== 第 {round_num} 輪：待查詢 {total_addresses} 個地址 =====', COLOR_YELLOW)}")

            # 按批次分組查詢
            for batch_num in range(0, total_addresses, BATCH_SIZE):
                batch_end = min(batch_num + BATCH_SIZE, total_addresses)
                batch_addresses = pending_addresses[batch_num:batch_end]
                batch_number = (batch_num // BATCH_SIZE) + 1
                total_batches = (total_addresses + BATCH_SIZE - 1) // BATCH_SIZE

                print(f"\n【第 {batch_number}/{total_batches} 批】查詢地址 {batch_num + 1}-{batch_end}")
                print("-" * 70)

                # 使用線程池並行查詢本批次的所有地址
                with ThreadPoolExecutor(max_workers=CONCURRENT_REQUESTS) as executor:
                    future_to_address = {
                        executor.submit(self.get_transaction_count, addr): (idx + batch_num + 1, addr)
                        for idx, addr in enumerate(batch_addresses)
                    }

                    for future in as_completed(future_to_address):
                        current_idx, address = future_to_address[future]
                        try:
                            tx_count, err_msg = future.result()
                            if tx_count is not None:
                                all_results[address] = tx_count
                                print_success(f"[{current_idx}/{total_addresses}] 查詢 {address[:10]}...{address[-8:]} → {tx_count} 筆交易")
                            else:
                                round_failures.append(address)
                                shown_err = err_msg or "查詢失敗"
                                print_error(f"[{current_idx}/{total_addresses}] 查詢 {address[:10]}...{address[-8:]} ✗ {shown_err}")
                        except Exception as e:
                            round_failures.append(address)
                            err_msg = str(e) if DEBUG_MODE else self._short_error_message(str(e))
                            print_error(f"[{current_idx}/{total_addresses}] 查詢 {address[:10]}...{address[-8:]} ✗ 異常: {err_msg}")

                # 如果還有下一批，顯示延遲信息
                if batch_end < total_addresses and BATCH_DELAY > 0:
                    print(f"\n⏳ 暫停 {BATCH_DELAY} 秒以避免RPC限制...")
                    time.sleep(BATCH_DELAY)

            # 不啟用「失敗地址分輪重試」則僅跑一輪
            if not RETRY_FAILED_WALLETS:
                break

            # 全部成功則結束
            if not round_failures:
                print_success(f"\n✓ 第 {round_num} 輪後已全部成功查詢。")
                break

            # 若未啟用直到全成功，則只跑第一輪
            if not RETRY_UNTIL_ALL_SUCCESS:
                print_warn(f"\n⚠️  仍有 {len(round_failures)} 個地址失敗，因配置限制不再重試。")
                break

            print_warn(f"\n⚠️  第 {round_num} 輪仍有 {len(round_failures)} 個地址失敗，下一輪僅重試失敗地址。")
            pending_addresses = round_failures
            round_num += 1

            if RETRY_ROUND_DELAY > 0:
                print(f"⏳ 輪次間等待 {RETRY_ROUND_DELAY} 秒...")
                time.sleep(RETRY_ROUND_DELAY)

        return all_results
    
    def print_summary(self, results: Dict[str, int]):
        """打印查詢結果摘要"""
        if not results:
            print("\n⚠️  沒有成功查詢的地址")
            return
        
        print("\n" + "="*70)
        print("查詢結果摘要")
        print("="*70)
        
        total_tx = 0
        
        for address, tx_count in results.items():
            print(f"{address} → {tx_count:>10} 筆交易")
            total_tx += tx_count
        
        print("="*70)
        print(f"{'總交易數':.<50} {total_tx} 筆")
        print(f"{'平均交易數 (每個地址)':.<50} {total_tx/len(results):.2f} 筆")
        print(f"{'查詢成功的地址數':.<50} {len(results)} 個")
        
        # 代理統計
        if self.use_proxy and self.proxy_manager and self.proxy_manager.proxies:
            print(f"{'總請求次數':.<50} {self.request_count} 次")
            print(f"{'代理錯誤次數':.<50} {self.proxy_error_count} 次")
        
        print("="*70)


def save_results(results: Dict[str, int], formats: List[str] = None):
    """
    以多種格式保存結果
    
    Args:
        results: 查詢結果字典 {地址: 交易數}
        formats: 輸出格式列表 ['json', 'csv', 'xlsx']
    """
    if formats is None:
        formats = OUTPUT_FORMATS
    
    if not results:
        print("\n⚠️  沒有結果可保存")
        return
    
    output_dir = Path("output")
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_base = f"tx_results_{timestamp}"

    print(f"\n💾 正在保存結果到 {output_dir}/ ...\n")
    
    for fmt in formats:
        try:
            if fmt.lower() == 'json':
                _save_json(results, str(output_dir / f"{filename_base}.json"))
            elif fmt.lower() == 'csv':
                _save_csv(results, str(output_dir / f"{filename_base}.csv"))
            elif fmt.lower() == 'xlsx':
                _save_xlsx(results, str(output_dir / f"{filename_base}.xlsx"))
            else:
                print(f"⚠️  未知的格式: {fmt}")
        except Exception as e:
            print(f"✗ 保存 {fmt.upper()} 失敗: {str(e)}")


def _save_json(results: Dict[str, int], filename: str):
    """保存為JSON格式"""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"✓ 已保存 JSON: {filename} ({len(results)} 筆記錄)")


def _save_csv(results: Dict[str, int], filename: str):
    """保存為CSV格式"""
    import csv
    
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        
        # 寫入標題
        writer.writerow(['地址', '交易數'])
        
        # 寫入數據
        for address, tx_count in results.items():
            writer.writerow([address, tx_count])
    
    print(f"✓ 已保存 CSV: {filename} ({len(results)} 筆記錄)")


def _save_xlsx(results: Dict[str, int], filename: str):
    """保存為Excel格式"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # 創建 workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '查詢結果'
        
        # 設置標題
        headers = ['JSON格式', '地址', '交易數']
        ws.append(headers)
        
        # 設置標題格式
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # 寫入數據
        for address, tx_count in results.items():
            json_format = f'{address}: {tx_count}'
            ws.append([json_format, address, tx_count])
        
        # 調整列寬
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        
        # 保存
        wb.save(filename)
        print(f"✓ 已保存 Excel: {filename} ({len(results)} 筆記錄)")
        
    except ImportError:
        print(f"✗ 保存 Excel 需要安裝: pip install openpyxl")


def load_addresses(filename: str = 'addresses.txt') -> List[str]:
    """
    從文件加載地址列表
    
    Args:
        filename: 地址文件名
        
    Returns:
        地址列表
    """
    addresses = []
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    addresses.append(line)
        
        if addresses:
            print(f"✓ 已從 {filename} 加載 {len(addresses)} 個地址\n")
        else:
            print(f"⚠️  {filename} 文件為空\n")
        return addresses
    
    except FileNotFoundError:
        print(f"✗ 找不到文件: {filename}")
        return []
    except Exception as e:
        print(f"✗ 讀取文件出錯: {str(e)}")
        return []


def main():
    """主程序"""
    print("🔍 區塊鏈交易查詢工具 (支持代理)")
    print("-" * 70)
    print(f"網絡: {SELECTED_NETWORK}")
    print(f"代理模式: {'✓ 已啟用' if USE_PROXY else '✗ 已禁用'}")
    print(f"輸出格式: {', '.join(OUTPUT_FORMATS)}")
    print(f"配置文件: config.toml")
    print("-" * 70 + "\n")
    
    if not ensure_file_with_sample(IMPORT_FILE):
        print(f"❌ 地址文件不可用，請檢查 {IMPORT_FILE}")
        return

    if USE_PROXY and not ensure_file_with_sample(PROXY_FILE):
        print(f"❌ 代理文件不可用，請檢查 {PROXY_FILE}")
        return

    # 加載地址
    addresses = load_addresses(IMPORT_FILE)
    
    if not addresses:
        print(f"❌ 沒有可查詢的地址，請檢查 {IMPORT_FILE} 文件")
        return
    
    # 初始化查詢器
    try:
        checker = TxChecker(network=SELECTED_NETWORK, use_proxy=USE_PROXY)
    except ValueError as e:
        print(f"❌ 初始化失敗: {str(e)}")
        return
    
    # 查詢交易數量
    results = checker.check_multiple_addresses(addresses)
    
    # 打印摘要
    checker.print_summary(results)
    
    # 保存結果到多種格式
    save_results(results, OUTPUT_FORMATS)


if __name__ == '__main__':
    # 禁用SSL警告（用於代理連接）
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    main()
